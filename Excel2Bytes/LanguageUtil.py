import os
import pandas as pd
from openpyxl import load_workbook

from LogUtil import ShowLog
from PathUtil import LanguageXlsxPath

IsUpdateLng = False
CNLanguage = 'cn'
LanguageKey = ['tw', 'en']
LanguageDict = {}


def InitLanguage():
    ShowLog('初始化语言表')
    if (not os.path.exists(LanguageXlsxPath) or
            CNLanguage not in pd.read_excel(LanguageXlsxPath, sheet_name=None).keys()):
        SaveLangData(CNLanguage, {'c': 'c', 'ID': 'text', 'uint': 'string', '编号': '文本'})
    ReadLangData(CNLanguage)
    if IsUpdateLng:
        for key in LanguageKey:
            ReadLangData(key)


def SaveLanguage():
    ShowLog('保存语言表')
    SaveLangData(CNLanguage, LanguageDict[CNLanguage])
    if IsUpdateLng:
        languageDict = {}
        for key, value in LanguageDict[CNLanguage].items():
            languageDict[key] = value
            for keyLng in LanguageKey:
                if key in LanguageDict[keyLng].keys():
                    languageDict[keyLng][key] = LanguageDict[keyLng][key]
                else:
                    languageDict[keyLng][key] = value
        for keyLng in LanguageKey:
            SaveLangData(keyLng, languageDict[keyLng])


def ReadLangData(key):
    ShowLog(f'读取语言表数据: {key}')
    # 读取Excel文件
    df = pd.read_excel(LanguageXlsxPath, sheet_name=None, header=None)
    # 判断是否包含名为key的页签
    if key not in df.keys():
        df = pd.DataFrame(columns=['c', 'c'])
        df.loc[1] = ['ID', 'text']
        df.loc[2] = ['uint', 'string']
        df.loc[3] = ['编号', '文本']
        lng = {}
        for key, value in LanguageDict[CNLanguage].items():
            df.loc[len(df)] = [key, value]
            lng[key] = value
        LanguageDict[key] = lng
        SaveLangData(key, lng)
    else:
        df = pd.read_excel(LanguageXlsxPath, sheet_name=key)
        keyLng = {}
        for index, row in df.iloc[3:].iterrows():
            keyLng[row.iloc[0]] = row.iloc[1]
            print(row.iloc[0], row.iloc[1])
        LanguageDict[key] = keyLng


def SaveLangData(key, data):
    ShowLog(f'保存语言表数据: {key}')
    workbook = load_workbook(filename=LanguageXlsxPath)
    worksheet = workbook.get_sheet_by_name(key)
    worksheet.delete_rows(5, worksheet.max_row)
    for key, value in data.items():
        worksheet.append([key, value])
    workbook.save(LanguageXlsxPath)


def GetLanguageKey(value):
    if value in LanguageDict[CNLanguage].values():
        for key, val in LanguageDict[CNLanguage].items():
            if val == value:
                return key
    else:
        if len(LanguageDict[CNLanguage]) == 0:
            LanguageDict[CNLanguage][1] = value
            return 1
        maxKey = max(LanguageDict[CNLanguage].keys())
        newKey = maxKey + 1
        LanguageDict[CNLanguage][newKey] = value
        if IsUpdateLng:
            for key in LanguageKey:
                LanguageDict[key][newKey] = value
        return newKey

